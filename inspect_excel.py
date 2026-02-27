
import pandas as pd
import openpyxl

file_path = "ITAC_Database.xlsx"

try:
    print(f"Opening {file_path}...")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    print("Sheet names:", wb.sheetnames)
    
    for sheet in wb.sheetnames:
        print(f"\n--- Sheet: {sheet} ---")
        ws = wb[sheet]
        # Get header row (assuming row 1)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        k = 0
        while all(h is None for h in headers) and k < 5:
             # Try next few rows if first is empty
             headers = [cell.value for cell in next(ws.iter_rows(min_row=k+2, max_row=k+2))]
             k += 1

        print("Columns (First non-empty row):", headers)
        
except Exception as e:
    print(f"Error inspecting file: {e}")
